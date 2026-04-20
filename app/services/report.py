"""
app/services/report.py

Hisobotlar:
  - Moliyaviy hisobot (oylik daromad, qarzlar)
  - Davomat hisoboti (guruh/o'quvchi bo'yicha)
  - O'qituvchi ish haqi
  - Qarzdorlar ro'yxati
  - Barcha o'quvchilar ro'yxati
  - Barcha o'qituvchilar ro'yxati

Format: Excel (openpyxl).
"""
import io
import uuid
from datetime import date
from typing import Optional

from sqlalchemy import and_, extract, func, select
from sqlalchemy.ext.asyncio import AsyncSession


# ─── Excel yordamchi ─────────────────────────────────────────────────

def _new_workbook(title: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl o'rnatilmagan: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = title

    HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    BORDER_SIDE = Side(style="thin", color="D1D5DB")
    CELL_BORDER = Border(
        left=BORDER_SIDE, right=BORDER_SIDE,
        top=BORDER_SIDE,  bottom=BORDER_SIDE,
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return wb, ws, {
        "header_fill": HEADER_FILL,
        "header_font": HEADER_FONT,
        "border":      CELL_BORDER,
        "center":      CENTER,
    }


def _set_headers(ws, styles: dict, headers: list, row: int = 1):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill      = styles["header_fill"]
        cell.font      = styles["header_font"]
        cell.border    = styles["border"]
        cell.alignment = styles["center"]
        ws.column_dimensions[cell.column_letter].width = max(len(str(header)) + 4, 14)


def _set_row(ws, styles: dict, values: list, row: int):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border    = styles["border"]
        cell.alignment = styles["center"]


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 1. Moliyaviy hisobot ─────────────────────────────────────────────

async def financial_report(
    db:        AsyncSession,
    month:     int,
    year:      int,
    branch_id: Optional[uuid.UUID] = None,
) -> bytes:
    """Oylik moliyaviy hisobot Excel."""
    from app.models.tenant import Payment, Student, User, Group

    stmt = (
        select(Payment, Student, User, Group)
        .join(Student, Payment.student_id == Student.id)
        .join(User,    Student.user_id    == User.id)
        .outerjoin(Group, Payment.group_id == Group.id)
        .where(
            and_(
                extract("month", Payment.created_at) == month,
                extract("year",  Payment.created_at) == year,
            )
        )
        .order_by(Payment.created_at.desc())
    )
    if branch_id:
        stmt = stmt.where(Student.branch_id == branch_id)

    rows = (await db.execute(stmt)).all()

    wb, ws, styles = _new_workbook(f"Moliyaviy hisobot {month}-{year}")
    from openpyxl.styles import Font

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"Moliyaviy hisobot — {month}/{year}"
    ws["A1"].font  = Font(bold=True, size=14)

    headers = ["#", "O'quvchi", "Guruh", "Summa (so'm)", "Usul", "Sana", "Holat"]
    _set_headers(ws, styles, headers, row=3)

    total_sum = 0
    for i, (pay, st, u, g) in enumerate(rows, start=1):
        _set_row(ws, styles, [
            i,
            f"{u.first_name} {u.last_name or ''}".strip(),
            g.name if g else "—",
            float(pay.amount),
            "Click" if pay.payment_method == "click" else "Naqd",
            pay.paid_at.strftime("%d.%m.%Y") if pay.paid_at else "—",
            "Tasdiqlandi" if pay.status == "completed" else pay.status,
        ], row=i + 3)
        if pay.status == "completed":
            total_sum += float(pay.amount)

    last = len(rows) + 5
    ws.cell(row=last, column=3, value="JAMI:").font = Font(bold=True)
    ws.cell(row=last, column=4, value=total_sum).font = Font(bold=True)

    return _to_bytes(wb)


# ─── 2. Davomat hisoboti ──────────────────────────────────────────────

async def attendance_report(
    db:        AsyncSession,
    month:     int,
    year:      int,
    group_id:  Optional[uuid.UUID] = None,
    branch_id: Optional[uuid.UUID] = None,
) -> bytes:
    """Oylik davomat hisoboti Excel."""
    from app.models.tenant import Attendance, Student, User, Group

    stmt = (
        select(
            User.first_name,
            User.last_name,
            Group.name.label("group_name"),
            func.count(Attendance.id).label("total"),
            func.count(Attendance.id).filter(Attendance.status == "present").label("present"),
            func.count(Attendance.id).filter(Attendance.status == "absent").label("absent"),
            func.count(Attendance.id).filter(Attendance.status == "late").label("late"),
            func.count(Attendance.id).filter(Attendance.status == "excused").label("excused"),
        )
        .select_from(Attendance)
        .join(Student, Attendance.student_id == Student.id)
        .join(User,    Student.user_id        == User.id)
        .join(Group,   Attendance.group_id    == Group.id)
        .where(
            and_(
                extract("month", Attendance.date) == month,
                extract("year",  Attendance.date) == year,
            )
        )
        .group_by(User.first_name, User.last_name, Group.name)
        .order_by(Group.name, User.first_name)
    )

    if group_id:
        stmt = stmt.where(Attendance.group_id == group_id)
    if branch_id:
        stmt = stmt.where(Group.branch_id == branch_id)

    rows = (await db.execute(stmt)).all()

    wb, ws, styles = _new_workbook(f"Davomat {month}-{year}")
    from openpyxl.styles import Font

    ws.merge_cells("A1:I1")
    ws["A1"].value = f"Davomat hisoboti — {month}/{year}"
    ws["A1"].font  = Font(bold=True, size=14)

    headers = ["#", "O'quvchi", "Guruh", "Jami", "Keldi", "Kelmadi", "Kechikdi", "Uzrli", "Foiz (%)"]
    _set_headers(ws, styles, headers, row=3)

    for i, row in enumerate(rows, start=1):
        total   = row.total   or 0
        present = row.present or 0
        late    = row.late    or 0
        pct     = round((present + late) / total * 100, 1) if total else 0
        _set_row(ws, styles, [
            i,
            f"{row.first_name} {row.last_name or ''}".strip(),
            row.group_name,
            total, present, row.absent or 0, late, row.excused or 0, pct,
        ], row=i + 3)

    return _to_bytes(wb)


# ─── 3. Qarzdorlar ro'yxati ───────────────────────────────────────────

async def debtors_report(
    db:        AsyncSession,
    branch_id: Optional[uuid.UUID] = None,
) -> bytes:
    """Qarzdorlar ro'yxati Excel."""
    from app.models.tenant import Student, User, StudentGroup, Group

    stmt = (
        select(Student, User)
        .join(User, Student.user_id == User.id)
        .where(and_(Student.balance < 0, Student.is_active == True))
        .order_by(Student.balance)
    )
    if branch_id:
        stmt = stmt.where(Student.branch_id == branch_id)

    rows = (await db.execute(stmt)).all()

    wb, ws, styles = _new_workbook("Qarzdorlar")
    from openpyxl.styles import Font

    ws["A1"].value = f"Qarzdorlar ro'yxati — {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font  = Font(bold=True, size=14)

    headers = ["#", "O'quvchi", "Telefon", "Qarz (so'm)", "Guruhlar"]
    _set_headers(ws, styles, headers, row=3)

    for i, (st, u) in enumerate(rows, start=1):
        g_stmt = (
            select(Group.name)
            .join(StudentGroup, StudentGroup.group_id == Group.id)
            .where(and_(StudentGroup.student_id == st.id, StudentGroup.is_active == True))
        )
        groups = ", ".join(r[0] for r in (await db.execute(g_stmt)).all())
        _set_row(ws, styles, [
            i,
            f"{u.first_name} {u.last_name or ''}".strip(),
            u.phone or "—",
            abs(float(st.balance)),
            groups or "—",
        ], row=i + 3)

    return _to_bytes(wb)


# ─── 4. O'qituvchi ish haqi ───────────────────────────────────────────

async def teacher_salary_report(
    db:        AsyncSession,
    month:     int,
    year:      int,
    branch_id: Optional[uuid.UUID] = None,
) -> bytes:
    """Barcha o'qituvchilar ish haqi hisoboti."""
    from app.models.tenant import Teacher, User, Group, Attendance

    stmt = (
        select(Teacher, User)
        .join(User, Teacher.user_id == User.id)
        .where(Teacher.is_active == True)
        .order_by(User.first_name)
    )
    if branch_id:
        stmt = stmt.where(Teacher.branch_id == branch_id)

    rows = (await db.execute(stmt)).all()

    wb, ws, styles = _new_workbook(f"Ish haqi {month}-{year}")
    from openpyxl.styles import Font

    ws["A1"].value = f"O'qituvchilar ish haqi — {month}/{year}"
    ws["A1"].font  = Font(bold=True, size=14)

    headers = ["#", "O'qituvchi", "Telefon", "Ish haqi turi", "Miqdor (so'm)",
               "Faol guruhlar", "Darslar soni", "Hisoblangan (so'm)"]
    _set_headers(ws, styles, headers, row=3)

    for i, (teacher, user) in enumerate(rows, start=1):
        g_cnt = (await db.execute(
            select(func.count(Group.id)).where(
                and_(Group.teacher_id == teacher.id, Group.status == "active")
            )
        )).scalar_one()

        l_cnt = (await db.execute(
            select(func.count(Attendance.id)).where(
                and_(
                    Attendance.teacher_id == teacher.id,
                    extract("month", Attendance.date) == month,
                    extract("year",  Attendance.date) == year,
                )
            )
        )).scalar_one()

        salary = float(teacher.salary_amount or 0)
        calculated = (
            salary        if teacher.salary_type == "fixed"
            else salary * l_cnt if teacher.salary_type == "per_lesson"
            else 0
        )

        _set_row(ws, styles, [
            i,
            f"{user.first_name} {user.last_name or ''}".strip(),
            user.phone or "—",
            teacher.salary_type or "—",
            salary,
            g_cnt,
            l_cnt,
            calculated,
        ], row=i + 3)

    return _to_bytes(wb)


# ─── 5. Barcha o'quvchilar ro'yxati ──────────────────────────────────

async def students_report(
    db:        AsyncSession,
    branch_id: Optional[uuid.UUID] = None,
) -> bytes:
    """Barcha faol o'quvchilar to'liq ma'lumotlari bilan."""
    from app.models.tenant import Student, User, StudentGroup, Group

    stmt = (
        select(Student, User)
        .join(User, Student.user_id == User.id)
        .where(Student.is_active == True)
        .order_by(User.first_name, User.last_name)
    )
    if branch_id:
        stmt = stmt.where(Student.branch_id == branch_id)

    rows = (await db.execute(stmt)).all()

    wb, ws, styles = _new_workbook("O'quvchilar")
    from openpyxl.styles import Font

    ws.merge_cells("A1:K1")
    ws["A1"].value = f"Barcha o'quvchilar ro'yxati — {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font  = Font(bold=True, size=14)

    headers = [
        "#", "Ism", "Familiya", "Telefon", "Email",
        "Tug'ilgan sana", "Jins", "Guruhlar",
        "Balans (so'm)", "Holat", "Qo'shilgan sana"
    ]
    _set_headers(ws, styles, headers, row=3)
    # Keng ustunlar
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["H"].width = 28

    for i, (st, u) in enumerate(rows, start=1):
        # Guruhlarni olish
        g_stmt = (
            select(Group.name)
            .join(StudentGroup, StudentGroup.group_id == Group.id)
            .where(and_(StudentGroup.student_id == st.id, StudentGroup.is_active == True))
        )
        groups = ", ".join(r[0] for r in (await db.execute(g_stmt)).all())

        gender_label = {"male": "Erkak", "female": "Ayol"}.get(st.gender or "", st.gender or "—")

        _set_row(ws, styles, [
            i,
            u.first_name,
            u.last_name  or "—",
            u.phone      or "—",
            u.email      or "—",
            st.date_of_birth.strftime("%d.%m.%Y") if st.date_of_birth else "—",
            gender_label,
            groups or "—",
            float(st.balance or 0),
            "Faol" if st.is_active else "Nofaol",
            st.created_at.strftime("%d.%m.%Y") if st.created_at else "—",
        ], row=i + 3)

    return _to_bytes(wb)


# ─── 6. Barcha o'qituvchilar ro'yxati ────────────────────────────────

async def teachers_report(
    db:        AsyncSession,
    branch_id: Optional[uuid.UUID] = None,
) -> bytes:
    """Barcha faol o'qituvchilar to'liq ma'lumotlari bilan."""
    from app.models.tenant import Teacher, User, Group

    stmt = (
        select(Teacher, User)
        .join(User, Teacher.user_id == User.id)
        .where(Teacher.is_active == True)
        .order_by(User.first_name, User.last_name)
    )
    if branch_id:
        stmt = stmt.where(Teacher.branch_id == branch_id)

    rows = (await db.execute(stmt)).all()

    wb, ws, styles = _new_workbook("O'qituvchilar")
    from openpyxl.styles import Font

    ws.merge_cells("A1:J1")
    ws["A1"].value = f"Barcha o'qituvchilar ro'yxati — {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font  = Font(bold=True, size=14)

    headers = [
        "#", "Ism", "Familiya", "Telefon", "Email",
        "Fanlar", "Ish haqi turi", "Ish haqi (so'm)",
        "Faol guruhlar", "Holat"
    ]
    _set_headers(ws, styles, headers, row=3)
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["F"].width = 28

    for i, (teacher, user) in enumerate(rows, start=1):
        g_cnt = (await db.execute(
            select(func.count(Group.id)).where(
                and_(Group.teacher_id == teacher.id, Group.status == "active")
            )
        )).scalar_one()

        _set_row(ws, styles, [
            i,
            user.first_name,
            user.last_name  or "—",
            user.phone      or "—",
            user.email      or "—",
            ", ".join(teacher.subjects or []) or "—",
            teacher.salary_type   or "—",
            float(teacher.salary_amount or 0),
            g_cnt,
            "Faol" if teacher.is_active else "Nofaol",
        ], row=i + 3)

    return _to_bytes(wb)
